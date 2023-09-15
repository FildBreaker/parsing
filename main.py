# Импорты
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Переменные для взаимодействия с Excel
workbook = Workbook()  # Создание новой рабочей книги
sheet = workbook.active  # Выбор активного листа

# Ссылка на сайт и данные, для захода на сайт как пользователь
url = "https://www.oat.ru/timetable/timetable/ul_volkhovstroya_5/%D0%9F%D0%AD321"
HEADERS = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36'
}

# Получение необходимых данных с сайта
response = requests.get(url, headers=HEADERS)
soup = bs(response.text, "lxml")

data = soup.find("section", class_="timetable-container")
table_rows = data.find_all("tr")

# Запись данных в таблицу Excel
for row_index, row in enumerate(table_rows, start=1):
    table_cells = row.find_all("td")
    for column_index, cell in enumerate(table_cells, start=1):
        sheet.cell(row=row_index, column=column_index).value = cell.text
        sheet.cell(row=row_index, column=column_index).alignment = Alignment(wrap_text=True)

# Выравнивание столбцов и строк по ширине и высоте текста
for column_cells in sheet.columns:
    max_length = 0
    for cell in column_cells:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

for row in sheet.rows:
    max_height = 0
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)
        if len(str(cell.value)) > max_height:
            max_height = len(str(cell.value))
    adjusted_height = (max_height + 2) * 4
    sheet.row_dimensions[row[0].row].height = adjusted_height

# Выравнивание итогового текста по центру
last_row = sheet.max_row
last_column = sheet.max_column
for column_cell in sheet.columns:
    sheet.cell(row=last_row, column=column_cell[0].column).alignment = Alignment(horizontal='center')

# Сохранение Excel
workbook.save(filename="data.xlsx")